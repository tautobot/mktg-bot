import os
import csv
from random import randint, choice
from datetime import datetime, timedelta

from datedelta import datedelta
from openpyxl import load_workbook
from pathlib import Path

from aqa.utils.enums import path

Path("/tmp/aqa").mkdir(parents=True, exist_ok=True)
filepath = os.path.dirname(__file__)
root_path = os.path.dirname(os.path.dirname(filepath))

def generate_nricfin(first=None, age=-1):
    numberList = ['S', 'T', 'F', 'G']
    if first is None:
        first = choice(numberList)
    if first != 'S' and first != 'T' and first != 'F' and first != 'G':
        return

    if not (age >= -1 and age <= 9):
        age = -1

    chars = list(str(randint(0, 9999999)).zfill(7))

    if age != -1: chars[0] = age
    output = first + ''.join(chars)

    chars[0] = str(int(chars[0]) * 2)
    chars[1] = str(int(chars[1]) * 7)
    chars[2] = str(int(chars[2]) * 6)
    chars[3] = str(int(chars[3]) * 5)
    chars[4] = str(int(chars[4]) * 4)
    chars[5] = str(int(chars[5]) * 3)
    chars[6] = str(int(chars[6]) * 2)

    sum = 0
    for i in range(7):
        sum += int(chars[i])

    offset = 4 if (first == "T" or first == "G") else 0
    temp = (offset + sum) % 11
    st = ["J", "Z", "I", "H", "G", "F", "E", "D", "C", "B", "A"]
    fg = ["X", "W", "U", "T", "R", "Q", "P", "N", "M", "L", "K"]

    if first == "S" or first == "T":
        alpha = st[temp]
    elif first == "F" or first == "G":
        alpha = fg[temp]
    else:
        alpha = "?"
    nric = output + alpha
    write_to_file(nric, '/tmp/aqa/nricfin')
    return nric

def generate_email():
    letters = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "q", "w", "e", "r", "t", "y", "u", "i", "o", "p", "0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
    name = ""
    for i in range(6):
        name = name + choice(letters)
    email = name + '@' + 'gigacover.com'
    write_to_file(email, '/tmp/aqa/email')
    return email

def generate_mobile_sgd():
    first = choice([855, 955, 811, 822, 833, 911, 922, 933])
    second = str(randint(1, 88888)).zfill(5)
    return f'{first}{second}'

def generate_mobile_phl():
    first = choice([907, 908, 909, 910, 919, 921, 922, 923, 928, 929, 932, 933])
    mid = str(randint(1, 999)).zfill(3)
    last = str(randint(1, 9999)).zfill(4)
    return f'{first}{mid}{last}'

def generate_nricfin_phl():
    nric = str(randint(1, 88888888888)).zfill(12)
    return nric

def cal_year(age):
    now = datetime.now()
    year = now.year - age
    return year

def cal_date(months, days):
    #TODO: Recheck with special dates
    # if months == -1 and datetime.now().strftime("%Y-%m-%d") in {'2020-03-31', '2020-05-31', '2020-07-31', '2020-10-31', '2020-12-31'}:
    #     days = days - 1
    # if months == 1 and datetime.now().strftime("%Y-%m-%d") in {'2020-01-31', '2020-03-31', '2020-05-31', '2020-08-31', '2020-10-31', '2020-12-31'}:
    #     days = days - 1
    # if months == -2 and datetime.now().strftime("%Y-%m-%d") in {'2020-04-29', '2020-08-30', '2020-01-31'}:
    #     days = days - 1
    # if months == -2 and datetime.now().strftime("%Y-%m-%d") in {'2020-04-30'}:
    #     days = days - 2
    # if months == 2 and datetime.now().strftime("%Y-%m-%d") in {'2020-12-31'}:
    #     days = days - 3
    # if months == 2 and datetime.now().strftime("%Y-%m-%d") in {'2020-12-30'}:
    #     days = days - 2
    # if months == 2 and datetime.now().strftime("%Y-%m-%d") in {'2020-12-29', '2020-07-31'}:
    #     days = days - 1
    return datetime.now() + datedelta(months=months, days=days)

def cal_date_by_date(date, months, days):
    return date + datedelta(months=months, days=days)

def write_to_file(name, path):
    with open(path, "w") as file:
        file.write(name)

def read_file(path):
    with open(path, "r") as file:
        return file.read()

def change(str):
    a = str.split('-')
    temp = a[0]
    a[0] = a[len(a) - 1]
    a[len(a) - 1] = temp
    a = '-'.join(a)
    return a

def change_datetime(date):
    a = datetime.strptime(date, "%Y-%m-%d")
    return a.strftime("%d %b %y").upper()

def cook_fixture_for_essentials():
    excel_file = f'{root_path}/aQA/fixtures/essentials_template.xlsx'
    nricfin = generate_nricfin()
    email = nricfin.lower() + '@gigacover.com'
    employee_no = 'AB-' + nricfin[1:-5] + '-' + nricfin[4:-1]

    wb = load_workbook(excel_file)
    ws = wb.active
    cell = ws[f'F2']; cell.value = email
    cell = ws[f'H2']; cell.value = employee_no
    cell = ws[f'D2']; cell.value = nricfin

    nricfin1 = generate_nricfin()
    email1 = nricfin1.lower() + '@gigacover.com'
    employee_no1 = 'AB-' + nricfin1[1:-5] + '-' + nricfin1[4:-1]
    cell = ws[f'F3']; cell.value = email1
    cell = ws[f'H3']; cell.value = employee_no1
    cell = ws[f'D3']; cell.value = nricfin1

    for i in range(4, 7):
        nricfin = generate_nricfin()
        email = nricfin.lower() + '@gigacover.com'
        employee_no = 'AB-' + nricfin[1:-5] + '-' + nricfin[4:-1]
        cell = ws[f'E{i}']; cell.value = nricfin
        cell = ws[f'D{i}']; cell.value = nricfin
        cell = ws[f'F{i}']; cell.value = email
        cell = ws[f'H{i}']; cell.value = employee_no
    wb.save(excel_file)
    return excel_file

def read_cell_in_excel_file(excel, cell):
    wb = load_workbook(excel, data_only=True)
    ws = wb.active
    cel = ws[cell].value
    return cel

def read_data_from_csv_file(csv_file, field_name):
    reader = csv.DictReader(open(csv_file))
    for raw in reader:
        return raw[field_name]

def cook_fixture_sponsor(file_name):
    excel_file = f'{path.fixture_dir}/{file_name}'

    wb = load_workbook(excel_file)
    ws = wb.active

    in_force = (datetime.now().date() - timedelta(days=1))
    expired = (datetime.now().date() - timedelta(days=32))
    paid = (datetime.now().date() + timedelta(days=32))

    nric_1 = generate_nricfin()
    nric_2 = generate_nricfin()
    nric_3 = generate_nricfin()
    nric_4 = generate_nricfin()
    nric_5 = generate_nricfin()
    nric_6 = generate_nricfin()
    nric_7 = generate_nricfin()

    ws['M2'].value = expired
    ws['M2'].number_format = 'dd/mm/yyyy'
    ws['G2'].value = nric_1
    ws['H2'].value = f'{nric_1}@gigacover.com'

    ws['M3'].value = in_force
    ws['M3'].number_format = 'dd/mm/yyyy'
    ws['G3'].value = nric_2
    ws['H3'].value = f'{nric_2}@gigacover.com'

    ws['M4'].value = paid
    ws['M4'].number_format = 'dd/mm/yyyy'
    ws['G4'].value = nric_3
    ws['H4'].value = f'{nric_3}@gigacover.com'

    ws['M5'].value = paid
    ws['M5'].number_format = 'dd/mm/yyyy'
    ws['G5'].value = nric_4
    ws['H5'].value = f'{nric_4}@gigacover.com'

    ws['M6'].value = expired
    ws['M6'].number_format = 'dd/mm/yyyy'
    ws['G6'].value = nric_5
    ws['H6'].value = f'{nric_5}@gigacover.com'

    ws['M7'].value = in_force
    ws['M7'].number_format = 'dd/mm/yyyy'
    ws['G7'].value = nric_6
    ws['H7'].value = f'{nric_6}@gigacover.com'

    ws['M8'].value = in_force
    ws['M8'].number_format = 'dd/mm/yyyy'
    ws['G8'].value = nric_7
    ws['H8'].value = f'{nric_7}@gigacover.com'

    wb.save(excel_file)
    return excel_file

def cook_fixture_bundle(file_name):
    excel_file = f'{path.fixture_dir}/{file_name}'

    wb = load_workbook(excel_file)
    ws = wb.active

    nric = generate_nricfin()
    email = f'{nric.lower()}@gigacover.com'
    in_force = datetime.now().date() - timedelta(days=1)

    ws['K2'].value = in_force
    ws['K2'].number_format = 'dd/mm/yyyy'
    ws['G2'].value = nric
    ws['H2'].value = email

    wb.save(excel_file)
    return excel_file

def cook_fixture_health(file_name):
    excel_file = f'{path.fixture_dir}/{file_name}'

    wb = load_workbook(excel_file)
    ws = wb.active

    primary_nric = generate_nricfin()
    dependent_nric = generate_nricfin()

    primary_email = f'{primary_nric.lower()}@gigacover.com'
    dependent_email = f'{dependent_nric.lower()}@gigacover.com'

    in_force = datetime.now().date() - timedelta(days=1)

    ws['O2'].value = in_force
    ws['O2'].number_format = 'dd/mm/yyyy'
    ws['H2'].value = primary_nric
    ws['K2'].value = primary_email

    ws['O3'].value = in_force
    ws['O3'].number_format = 'dd/mm/yyyy'
    ws['H3'].value = dependent_nric
    ws['K3'].value = dependent_email
    ws['R3'].value = primary_email

    wb.save(excel_file)
    return excel_file

def cook_fixture_pml():
    excel_file = f'{root_path}/aQA/fixtures/pml_template.xlsx'

    wb = load_workbook(excel_file)
    ws = wb.active

    nric1 = generate_nricfin()
    email1 = f'{nric1}@gigacover.com'

    nric2 = generate_nricfin()
    email2 = f'{nric2}@gigacover.com'
    in_force = datetime.now().date() - timedelta(days=1)

    ws['M2'].value = in_force
    ws['M2'].number_format = 'dd/mm/yyyy'
    ws['H2'].value = nric1
    ws['I2'].value = email1

    ws['M3'].value = in_force
    ws['M2'].number_format = 'dd/mm/yyyy'
    ws['H3'].value = nric2
    ws['I3'].value = email2

    wb.save(excel_file)
    return excel_file

def cook_fixture_eb():
    excel_file = f'{root_path}/aQA/fixtures/eb_template.xlsx'

    wb = load_workbook(excel_file)
    ws = wb.active

    primary_nric = generate_nricfin()
    primary_email = f'{primary_nric}@gigacover.com'

    dependent_nric = generate_nricfin()
    dependent_email = f'{dependent_nric}@gigacover.com'
    today = datetime.now().date()

    ws['L2'].number_format = 'dd/mm/yyyy'
    ws['L2'].value = today
    ws['D2'].value = primary_nric
    ws['G2'].value = primary_email

    ws['L2'].number_format = 'dd/mm/yyyy'
    ws['L3'].value = today
    ws['D3'].value = dependent_nric
    ws['G3'].value = dependent_email
    ws['O3'].value = primary_email

    wb.save(excel_file)
    return excel_file

def cook_fixture_mer():
    excel_file = f'{root_path}/aQA/fixtures/mer_template.xlsx'

    wb = load_workbook(excel_file)
    ws = wb.active

    vehicle_reg1 = generate_nricfin()
    vehicle_reg2 = generate_nricfin()
    paid = datetime.now().date() + timedelta(days=1)

    ws['N2'].number_format = 'dd/mm/yyyy'
    ws['N2'].value = paid
    ws['A2'].value = vehicle_reg1
    ws['E2'].value = f'{vehicle_reg1}@gigacover.com'

    ws['N3'].number_format = 'dd/mm/yyyy'
    ws['N3'].value = paid
    ws['A3'].value = vehicle_reg2
    ws['E3'].value = f'{vehicle_reg2}@gigacover.com'

    wb.save(excel_file)
    return excel_file

def cook_fixture_validation_eb():
    excel_file = f'{root_path}/aQA/fixtures/eb_validation_template.xlsx'

    tomorrow = datetime.now().date() + timedelta(days=1)
    yesterday = datetime.now().date() - timedelta(days=1)

    wb = load_workbook(excel_file)
    ws = wb.active

    ws['D2'].value = nric = generate_nricfin()
    ws['G2'].value = f'{nric}@gigacover.com'

    ws['D3'].value = nric = generate_nricfin()
    ws['G3'].value = f'{nric}@gigacover.com'
    ws['L3'].number_format = 'dd/mm/yyyy'
    ws['L3'].value = tomorrow

    ws['D4'].value = nric = generate_nricfin()
    ws['G4'].value = f'{nric}@gigacover.com'
    ws['L4'].number_format = 'dd/mm/yyyy'
    ws['L4'].value = tomorrow

    ws['D5'].value = nric = generate_nricfin()
    ws['G5'].value = f'{nric}@gigacover.com'
    ws['L5'].number_format = 'dd/mm/yyyy'
    ws['L5'].value = tomorrow

    ws['D6'].value = nric = generate_nricfin()
    ws['G6'].value = f'{nric}@gigacover.com'
    ws['L6'].number_format = 'dd/mm/yyyy'
    ws['L6'].value = yesterday

    ws['D7'].value = generate_nricfin()
    ws['L7'].number_format = 'dd/mm/yyyy'
    ws['L7'].value = tomorrow

    ws['D8'].value = nric = generate_nricfin()
    ws['G8'].value = f'{nric}@gigacover.com'
    ws['L8'].number_format = 'dd/mm/yyyy'
    ws['L8'].value = tomorrow

    ws['L9'].value = tomorrow

    wb.save(excel_file)
    return excel_file

def cook_fixture_parcel_validation():
    today = datetime.now().strftime("%d/%m/%Y")
    future_date = (datetime.now() + timedelta(days=1)).strftime("%d/%m/%Y")
    past_date = (datetime.now() - timedelta(days=1)).strftime("%d/%m/%Y")
    header = ['Parcel Reference Number','Item Name','Item Description','Purchase Date','Delivery Date','Item Cost','Courier/delivery vehicle details']
    missing_reference_number = ['','Iphone 13','test onboard parcel',today,today,'1000000','car']
    missing_item_name = ['Validation-00000','','test onboard parcel',today,today,'1000000','car']
    missing_item_description = ['Validation-00001','Iphone 13','',today,today,'1000000','car']
    missing_purchase_date = ['Validation-00002','Iphone 13','test onboard parcel','',today,'1000000','car']
    missing_delivery_date = ['Validation-00003','Iphone 13','test onboard parcel',today,'','1000000','car']
    missing_item_cost = ['Validation-00004','Iphone 13','test onboard parcel',today,today,'','car']
    missing_vehicle = ['Validation-00005','Iphone 13','test onboard parcel',today,today,'1000000','']
    invalid_purchase_date = ['Validation-00006','Iphone 13','test onboard parcel','2',today,'1000000','car']
    invalid_delivery_date = ['Validation-00007','Iphone 13','test onboard parcel',today,'2','1000000','car']
    purchase_date_after_delivery_date = ['Validation-00008','Iphone 13','test onboard parcel',future_date,today,'1000000','car']
    delivery_date_in_past = ['Validation-00009','Iphone 13','test onboard parcel',today,past_date,'1000000','car']
    valid_data = ['Validation-00010','Iphone 13','test onboard parcel',today,today,'1000000','car']
    existing_reference_number = ['Validation-00010','Iphone 13','test onboard parcel',today,today,'1000000','car']

    csv_file = f'{root_path}/aQA/fixtures/parcel_validation_template.csv'
    with open(csv_file, 'w', encoding='UTF8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerow(missing_reference_number)
        writer.writerow(missing_item_name)
        writer.writerow(missing_item_description)
        writer.writerow(missing_purchase_date)
        writer.writerow(missing_delivery_date)
        writer.writerow(missing_item_cost)
        writer.writerow(missing_vehicle)
        writer.writerow(invalid_purchase_date)
        writer.writerow(invalid_delivery_date)
        writer.writerow(purchase_date_after_delivery_date)
        writer.writerow(delivery_date_in_past)
        writer.writerow(valid_data)
        writer.writerow(existing_reference_number)

    return csv_file

def cook_fixture_parcel_valid_data():
    today = datetime.now().strftime("%d/%m/%Y")
    header = ['Parcel Reference Number','Item Name','Item Description','Purchase Date','Delivery Date','Item Cost','Courier/delivery vehicle details']
    valid_data_1 = ['Valid-00000','Iphone 13','test onboard parcel 1',today,today,'1000000','car']
    valid_data_2 = ['Valid-00001','Iphone 13','test onboard parcel 2',today,today,'1000000','car']
    valid_data_3 = ['Valid-00002','Iphone 13','test onboard parcel 3',today,today,'1000000','car']
    valid_data_4 = ['Valid-00003','Iphone 13','test onboard parcel 4',today,today,'1000000','car']

    csv_file = f'{root_path}/aQA/fixtures/parcel_upload_template.csv'
    with open(csv_file, 'w', encoding='UTF8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerow(valid_data_1)
        writer.writerow(valid_data_2)
        writer.writerow(valid_data_3)
        writer.writerow(valid_data_4)

    return csv_file
