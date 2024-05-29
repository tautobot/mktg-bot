import sys
import os
import openpyxl
from aqa.utils.generic import *

filepath = os.path.dirname(__file__)
sys.path.append(filepath)


def get_start_unique_nricfin(pnum, product):
    with open(f'{filepath}/unique_nricfin', 'a') as f:
        f.write(datetime.now().strftime(f'E%y%m%d{pnum}' + ('1' if product == 'flep' else '0')) + '\n')
        f.write(datetime.now().strftime(f'I%y%m%d{pnum}' + ('1' if product == 'flip' else '0')) + '\n')
        f.write(datetime.now().strftime(f'P%y%m%d{pnum}' + ('1' if product == 'pa' else '0')))
        if product == 'flep':
            return datetime.now().strftime(f'E%y%m%d{pnum}1')
        if product == 'flip':
            return datetime.now().strftime(f'I%y%m%d{pnum}1')
        if product == 'pa':
            return datetime.now().strftime(f'P%y%m%d{pnum}1')


def get_unique_nricfin(server, product):
    if server is "autoarmour@release" or "jarvis@staging5":
        pnum = '00'
    else:
        pnum = '000'
    existing_file = False
    if os.path.isfile(f'{filepath}/unique_nricfin'):
        existing_file = True

    if not existing_file:
        return get_start_unique_nricfin(pnum, product)
    else:
        unique_nricfin = None
        with open(f'{filepath}/unique_nricfin', 'r') as f:
            data = f.readlines()
            if data[0][1:7] != datetime.now().strftime('%y%m%d'):
                with open(f'{filepath}/unique_nricfin', 'w') as cleared_file:
                    cleared_file.close()
                return get_start_unique_nricfin(pnum, product)
            else:
                if product == 'flep':
                    seq_nric_num = int(data[0][7:]) + 1
                    unique_nricfin = datetime.now().strftime('E%y%m%d') + str(seq_nric_num).zfill(3) if server is "autoarmour@release" or "jarvis@staging5" else str(seq_nric_num).zfill(4)
                    data[0] = unique_nricfin + '\n'
                elif product == 'flip':
                    seq_nric_num = int(data[1][7:]) + 1
                    unique_nricfin = datetime.now().strftime('I%y%m%d') + str(seq_nric_num).zfill(3) if server is "autoarmour@release" or "jarvis@staging5" else str(seq_nric_num).zfill(4)
                    data[1] = unique_nricfin + '\n'
                elif product == 'pa':
                    seq_nric_num = int(data[2][7:]) + 1
                    unique_nricfin = datetime.now().strftime('P%y%m%d') + str(seq_nric_num).zfill(3) if server is "autoarmour@release" or "jarvis@staging5" else str(seq_nric_num).zfill(4)
                    data[2] = unique_nricfin

                with open(f'{filepath}/unique_nricfin', 'w') as file:
                    file.writelines(data)
                    file.close()
        return unique_nricfin


def get_product_from_file_name(file):
    if 'flep' in file:
        return 'flep'
    elif 'flip' in file:
        return 'flip'
    elif '_pa_' in file:
        return 'pa'
    elif '_gpac_' in file:
        return 'gpac'
    else:
        return None

def create_sp_file_all_columns(file, policy_start, status, row=2, nricfin=None):
    product = get_product_from_file_name(file)
    if nricfin is None:
        nricfin = get_unique_nricfin(product)

    email = nricfin.lower() + '@gigacover.com'
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # product B2 | nricfin G2 | email H2 | policy start L2
    ws[f'A{row}'] = 'GOJEK GoalBetter'
    ws[f'B{row}'] = product
    ws[f'C{row}'] = 'flep80'
    ws[f'D{row}'] = 'Monthly'
    ws[f'E{row}'] = 'Trieu'
    ws[f'F{row}'] = 'Truong'
    ws[f'G{row}'] = nricfin
    ws[f'H{row}'] = email
    ws[f'I{row}'] = '03/02/1984'
    ws[f'I{row}'].number_format = 'DD/MM/YYYY'
    ws[f'J{row}'] = '0938846797'
    ws[f'K{row}'] = '123456'
    ws[f'L{row}'] = policy_start
    ws[f'L{row}'].number_format = 'DD/MM/YYYY'
    ws[f'M{row}'] = 'driver'
    ws[f'N{row}'] = 2
    ws[f'O{row}'] = 'Y'
    ws[f'P{row}'] = 'N'
    ws[f'Q{row}'] = 'N'

    wb.save(file)
    write_data_log(file.split('testfiles/')[1], product, nricfin, policy_start.strftime('%Y-%m-%d'), status)
    return product, nricfin, email, policy_start

def create_sp_file(file, policy_start, status, row=2, nricfin=None):
    product = get_product_from_file_name(file)
    nricfin = generate_nricfin()

    # if nricfin is None:
    #     nricfin = get_unique_nricfin(server, product)

    email = nricfin + '@gigacover.com'
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    if product == "gpac":
        # policy number B2 | product C2 | nricfin H2 | email I2 | policy start M2
        ws[f'B{row}'] = 'NTUC' + datetime.now().strftime('%Y%m%d')
        ws[f'C{row}'] = product
        ws[f'H{row}'] = nricfin
        ws[f'I{row}'] = email
        ws[f'M{row}'] = policy_start
        ws[f'M{row}'].number_format = 'DD/MM/YYYY'
    else:
        # product B2 | nricfin G2 | email H2 | policy start L2
        ws[f'B{row}'] = product
        ws[f'G{row}'] = nricfin
        ws[f'H{row}'] = email
        ws[f'L{row}'] = policy_start
        ws[f'L{row}'].number_format = 'DD/MM/YYYY'

    wb.save(file)
    write_data_log(file.split('testfiles/')[1], product, nricfin, policy_start.strftime('%Y-%m-%d'), status)
    return product, nricfin, email, policy_start


def create_sp_from_existing(file, product, nricfin, email, policy_start, status):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # product B2 | nricfin G2 | email H2 | policy start L2
    ws['B2'] = product
    ws['G2'] = nricfin
    ws['H2'] = email
    ws['L2'] = policy_start
    ws['L2'].number_format = 'DD/MM/YYYY'
    wb.save(file)
    write_data_log(file.split('testfiles/')[1], product, nricfin, policy_start.strftime('%Y-%m-%d'), status)


# Create sponsored purchase file with some policies same nricfin, different policy start from array
def create_sps_file(file, arr_policy_start, status):
    product = get_product_from_file_name(file)
    nricfin = get_unique_nricfin(product)
    email = nricfin + '@gigacover.com'
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    start_row = 2
    policy_start_list = None
    for policy_start in arr_policy_start:
        # product B2 | nricfin G2 | email H2 | policy start L2
        ws[f'B{start_row}'] = product
        ws[f'G{start_row}'] = nricfin
        ws[f'H{start_row}'] = email
        ws[f'L{start_row}'] = policy_start
        ws[f'L{start_row}'].number_format = 'DD/MM/YYYY'
        policy_start_list = f"{policy_start.strftime('%Y-%m-%d')}" if policy_start_list is None else f"{policy_start_list}_{policy_start.strftime('%Y-%m-%d')}"
        start_row = start_row + 1
    wb.save(file)
    write_data_log(file.split('testfiles/')[1], product, nricfin, policy_start_list.strftime('%Y-%m-%d'), status)


def create_sp_file_for_validation(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    product = None
    if 'gpacs_id' in file:
        product = 'gpac_id'
    elif 'gpacs_sg' in file:
        product = 'gpac_sg'
    elif 'hcps_id' in file:
        product = 'hcp_id'
    existing_nric = generate_nricfin()
    existing_email = f'{existing_nric}@gigacover.com'
    existing_mobile = f'100-{datetime.now().strftime("%y%m%d%H%M%S")}-27'
    if product == ('gpac_id' or 'gpac_sg'):
        for row in range(2, 23):
            nricfin = generate_nricfin()
            email = f'{nricfin}@gigacover.com'
            # policy number B | product C | nricfin H | email I | DOB J
            if row == 13:
                ws[f'H{row}'] = ''
            else:
                ws[f'H{row}'] = nricfin
            if row == 14:
                ws[f'I{row}'] = ''
            else:
                ws[f'I{row}'] = email
            if row == 16:
                # DOB < 16 ages
                ws[f'J{row}'] = datetime.now() + datedelta(years=-16, months=0, days=1)
                ws[f'J{row}'].number_format = 'DD/MM/YYYY'
            if row == 17:
                # DOB > 70 ages
                ws[f'J{row}'] = datetime.now() + datedelta(years=-70, months=0, days=0)
                ws[f'J{row}'].number_format = 'DD/MM/YYYY'
            file_name = file.split('testfiles/')[1]
            addin_row_to_file_name = file_name[:8] + f'_{row}' + file_name[8:]
            write_data_log(addin_row_to_file_name, product, nricfin, 'null', 'null,FAIL')
    elif product == 'hcp_id':
        for row in range(2, 30):
            nricfin = generate_nricfin()
            email = f'{nricfin}@gigacover.com'
            # policy number N | product C | nricfin H | email I | DOB J
            if row == 3:
                ws[f'N{row}'] = ''
            else:
                ws[f'N{row}'] = f'HCP-GC-{nricfin}'
            if row == 13:
                ws[f'G{row}'] = ''
            elif row == 27:
                ws[f'G{row}'] = existing_nric
            else:
                ws[f'G{row}'] = nricfin

            if row == 14:
                ws[f'H{row}'] = ''
            elif row == 27:
                ws[f'H{row}'] = existing_email
                ws[f'J{row}'] = existing_mobile
            elif row == 28:
                ws[f'H{row}'] = existing_email
            else:
                ws[f'H{row}'] = email
            if row == 16:
                # DOB < 16 ages
                ws[f'I{row}'] = datetime.now() + datedelta(years=-16, months=0, days=1)
                ws[f'I{row}'].number_format = 'DD/MM/YYYY'
            if row == 17:
                # DOB > 75 ages
                ws[f'I{row}'] = datetime.now() + datedelta(years=-75, months=0, days=0)
                ws[f'I{row}'].number_format = 'DD/MM/YYYY'
            if row == 18:
                ws[f'J{row}'] = ''
            elif row == 29:
                ws[f'J{row}'] = existing_mobile
            else:
                ws[f'J{row}'] = f'100-{datetime.now().strftime("%y%m%d%H%M%S")}-{str(row).zfill(2)}'

            if row == 20:
                ws[f'O{row}'] = ''
            else:
                ws[f'O{row}'] = f'HCP-AXA-{nricfin}'
            file_name = file.split('testfiles/')[1]
            addin_row_to_file_name = file_name[:8] + f'_{row}' + file_name[8:]
            write_data_log(addin_row_to_file_name, product, nricfin, 'null', 'null,null')

    wb.save(file)



def create_sp_import_file(file, nricfin, email, policy_start, status, row=2):
    product = get_product_from_file_name(file)
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    if product == "gpac":
        # policy number B2 | product C2 | nricfin H2 | email I2 | policy start M2
        ws[f'B{row}'] = "NTUC" + datetime.now().strftime('%Y%m%d')
        ws[f'C{row}'] = product
        ws[f'H{row}'] = nricfin
        ws[f'I{row}'] = email
        ws[f'M{row}'] = policy_start
        ws[f'M{row}'].number_format = 'DD/MM/YYYY'
    else:
        # product B2 | nricfin G2 | email H2 | policy start L2
        ws[f'B{row}'] = product
        ws[f'G{row}'] = nricfin
        ws[f'H{row}'] = email
        ws[f'L{row}'] = policy_start
        ws[f'L{row}'].number_format = 'DD/MM/YYYY'

    wb.save(file)
    write_data_log(file.split('testfiles/')[1], product, nricfin, policy_start.strftime('%Y-%m-%d'), status)
    return product, nricfin, email, policy_start


# def create_onboard_benefits_file_all_columns(file, policy_start, emp_no=1, row=2, nricfin=None):
#     if nricfin is None:
#         nricfin = generate_nricfin()
#
#     email = nricfin.lower() + '@gigacover.com'
#     wb = openpyxl.load_workbook(file)
#     ws = wb.active
#
#     # product B2 | nricfin G2 | email H2 | policy start L2
#     ws[f'A{row}'] = 'GIG'
#     ws[f'B{row}'] = 'Gigacover'
#     ws[f'C{row}'] = 'Life'
#     ws[f'D{row}'] = f'Trieu {nricfin}'
#     ws[f'E{row}'] = nricfin
#     ws[f'F{row}'] = email
#     ws[f'G{row}'] = 'male'
#     ws[f'H{row}'] = 'AB-100-' + str(emp_no).zfill(2)
#     ws[f'I{row}'] = 'Primary'
#     ws[f'J{row}'] = 'N/A'
#     ws[f'K{row}'] = '17/05/1983'
#     ws[f'K{row}'].number_format = 'DD/MM/YYYY'
#     ws[f'L{row}'] = '0938846797'
#     ws[f'M{row}'] = '123456'
#     ws[f'N{row}'] = policy_start
#     ws[f'O{row}'].number_format = 'DD/MM/YYYY'
#     ws[f'P{row}'] = 'driver'
#     ws[f'Q{row}'] = 2
#     ws[f'R{row}'] = 'Y'
#     ws[f'S{row}'] = 'N'
#     ws[f'T{row}'] = 'N'
#
#     wb.save(file)
#     write_data_log(file.split('testfiles/')[1], product, nricfin, policy_start.strftime('%Y-%m-%d'), status)
#     return product, nricfin, email, policy_start


def write_data_log(case, product, nricfin, policy_start, status):
    if 'import' in case:
        file = f'{filepath}/testfiles/import_data_log'
    elif 'validation' in case:
        file = f'{filepath}/testfiles/validation_data_log'
    else:
        file = f'{filepath}/testfiles/existing_data_log'
    caseid = case.split("_auto_")[0]
    # if 'import' in case:
    #     caseid = case[:8]
    # elif 'validation' in case:
    #     caseid = case[:10]
    # elif 'temp' in case:
    #     caseid = case[:7]
    # else:
    #     caseid = case[:6]

    existing_file = False
    if os.path.isfile(file):
        existing_file = True

    if not existing_file:
        with open(file, 'a+') as a:
            a.write(f'{caseid}:{product},{nricfin},{policy_start},{status}' + '\n')
    else:
        with open(file, 'r') as f:
            i = 0
            existing_data_case = False
            lines = f.readlines()
            for line in lines:
                if line.split(':')[0] == caseid:
                    existing_data_case = True
                    with open(file, 'w') as w:
                        lines[i] = f'{caseid}:{product},{nricfin},{policy_start},{status}' + '\n'
                        w.writelines(lines)
                        w.close()
                        break
                i = i + 1
            if not existing_data_case:
                with open(file, 'a') as a:
                    a.write(f'{caseid}:{product},{nricfin},{policy_start},{status}' + '\n')


